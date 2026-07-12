"""Geracao do relatorio Excel (linha do tempo + resumos agregados)."""

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import classificar_alarmes
import classificar_bypass
from sessoes import formatar_duracao


AZUL = "2E75B6"
BRANCO = "FFFFFF"
LINHA_PAR = "EBF3FB"
VERMELHO_BG = "FFC7CE"
VERMELHO_FG = "9C0006"

# Objetos de estilo compartilhados: criar um novo Font/Fill/Border a cada
# celula (openpyxl) fica muito lento em planilhas com dezenas de milhares de
# linhas (ex: ~65 mil linhas na aba Linha do Tempo levavam ~90s). Reaproveitar
# as mesmas instancias reduz isso para poucos segundos.
_LADO = Side(style="thin", color="BFBFBF")
_BORDA = Border(left=_LADO, right=_LADO, top=_LADO, bottom=_LADO)
_ALINHAMENTO_CENTRO = Alignment(horizontal="center", vertical="center")
_ALINHAMENTO_ESQUERDA = Alignment(horizontal="left", vertical="center")

_FONTE_CABECALHO = Font(bold=True, size=10, color=BRANCO)
_FUNDO_CABECALHO = PatternFill("solid", fgColor=AZUL)

_FONTE_NORMAL = Font(size=9)
_FONTE_DESTAQUE = Font(size=9, color=VERMELHO_FG, bold=True)
_FUNDO_PAR = PatternFill("solid", fgColor=LINHA_PAR)
_FUNDO_IMPAR = PatternFill("solid", fgColor=BRANCO)
_FUNDO_DESTAQUE = PatternFill("solid", fgColor=VERMELHO_BG)


def _hcell(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = _FONTE_CABECALHO
    cell.fill = _FUNDO_CABECALHO
    cell.alignment = _ALINHAMENTO_CENTRO
    cell.border = _BORDA
    return cell


def _dcell(ws, r, c, v, destaque=False):
    cell = ws.cell(row=r, column=c, value=v)
    if destaque:
        cell.fill = _FUNDO_DESTAQUE
        cell.font = _FONTE_DESTAQUE
    else:
        cell.fill = _FUNDO_PAR if r % 2 == 0 else _FUNDO_IMPAR
        cell.font = _FONTE_NORMAL
    cell.alignment = _ALINHAMENTO_ESQUERDA
    cell.border = _BORDA
    return cell


def _autoajustar_colunas(ws, larguras):
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura


def _adicionar_tabela(ws, nome, num_linhas, num_colunas):
    ref = f"A1:{get_column_letter(num_colunas)}{num_linhas}"
    tabela = Table(displayName=nome, ref=ref)
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(tabela)


ACOES_ALARME = {
    "UNACK": "Atuou (nao reconhecido)",
    "ACK": "Reconhecido pelo operador",
    "ACK_RTN": "Retornou ao normal (ja reconhecido)",
}


def _linha_do_tempo(eventos):
    linhas = []
    for e in eventos:
        if classificar_alarmes.eh_evento_de_alarme(e):
            linhas.append({
                "dt": e["dt"], "tipo": "Alarme", "tag": e["tag"],
                "descricao": e["descricao"], "operador": e["operador"],
                "acao": ACOES_ALARME.get(e["tipo_evento"], e["tipo_evento"]),
                "estado_novo": e["estado_novo"], "estado_anterior": e["estado_anterior"],
            })
        elif classificar_bypass.eh_evento_de_bypass(e) and e["tipo_evento"] == "EVT":
            sinal = classificar_bypass.sinal_evento_w(e)
            if sinal:
                linhas.append({
                    "dt": e["dt"], "tipo": "Bypass", "tag": e["tag"],
                    "descricao": e["descricao"], "operador": e["operador"],
                    "acao": "Ativado" if sinal == "ON" else "Removido",
                    "estado_novo": e["estado_novo"], "estado_anterior": e["estado_anterior"],
                })
    linhas.sort(key=lambda l: l["dt"])
    return linhas


def _aba_linha_do_tempo(wb, eventos):
    ws = wb.create_sheet("Linha do Tempo")
    cabecalho = ["Data", "Hora", "Tipo", "Tag", "Descricao", "Operador", "Acao", "Estado Novo", "Estado Anterior"]
    for c, titulo in enumerate(cabecalho, start=1):
        _hcell(ws, 1, c, titulo)

    linhas = _linha_do_tempo(eventos)
    for i, l in enumerate(linhas, start=2):
        destaque = l["tipo"] == "Bypass"
        _dcell(ws, i, 1, l["dt"].strftime("%d/%m/%Y"), destaque)
        _dcell(ws, i, 2, l["dt"].strftime("%H:%M:%S"), destaque)
        _dcell(ws, i, 3, l["tipo"], destaque)
        _dcell(ws, i, 4, l["tag"], destaque)
        _dcell(ws, i, 5, l["descricao"], destaque)
        _dcell(ws, i, 6, l["operador"], destaque)
        _dcell(ws, i, 7, l["acao"], destaque)
        _dcell(ws, i, 8, l["estado_novo"], destaque)
        _dcell(ws, i, 9, l["estado_anterior"], destaque)

    _autoajustar_colunas(ws, [12, 11, 9, 28, 40, 12, 32, 16, 16])
    ws.freeze_panes = "A2"
    if linhas:
        _adicionar_tabela(ws, "LinhaDoTempo", len(linhas) + 1, len(cabecalho))


def _aba_resumo_alarmes(wb, resumo_alarmes):
    ws = wb.create_sheet("Resumo Alarmes")
    cabecalho = [
        "Tag", "Descricao", "Nº Atuacoes", "Retornou Sozinho", "Reconhecidas",
        "Ainda Ativo", "Tempo Total Atuado", "Tempo Medio ate Reconhecimento",
    ]
    for c, titulo in enumerate(cabecalho, start=1):
        _hcell(ws, 1, c, titulo)

    resumo_ordenado = sorted(resumo_alarmes, key=lambda r: -r["tempo_total_atuado_s"])
    for i, r in enumerate(resumo_ordenado, start=2):
        destaque = r["qtd_ainda_ativo"] > 0
        _dcell(ws, i, 1, r["tag"], destaque)
        _dcell(ws, i, 2, r["descricao"], destaque)
        _dcell(ws, i, 3, r["qtd_atuacoes"], destaque)
        _dcell(ws, i, 4, r["qtd_retornou_sozinho"], destaque)
        _dcell(ws, i, 5, r["qtd_reconhecidas"], destaque)
        _dcell(ws, i, 6, r["qtd_ainda_ativo"], destaque)
        _dcell(ws, i, 7, r["tempo_total_atuado_fmt"], destaque)
        _dcell(ws, i, 8, r["tempo_medio_reconhecimento_fmt"], destaque)

    _autoajustar_colunas(ws, [26, 40, 12, 16, 14, 12, 18, 26])
    ws.freeze_panes = "A2"
    if resumo_ordenado:
        _adicionar_tabela(ws, "ResumoAlarmes", len(resumo_ordenado) + 1, len(cabecalho))


def _aba_resumo_bypass(wb, resumo_bypass):
    ws = wb.create_sheet("Resumo Bypass")
    cabecalho = [
        "Tag", "Descricao", "Nº Ativacoes", "Ainda em Bypass",
        "Tempo Total em Bypass", "Operadores Responsaveis",
    ]
    for c, titulo in enumerate(cabecalho, start=1):
        _hcell(ws, 1, c, titulo)

    resumo_ordenado = sorted(resumo_bypass, key=lambda r: -r["tempo_total_bypass_s"])
    for i, r in enumerate(resumo_ordenado, start=2):
        destaque = r["qtd_ainda_ativo"] > 0
        _dcell(ws, i, 1, r["tag"], destaque)
        _dcell(ws, i, 2, r["descricao"], destaque)
        _dcell(ws, i, 3, r["qtd_ativacoes"], destaque)
        _dcell(ws, i, 4, r["qtd_ainda_ativo"], destaque)
        _dcell(ws, i, 5, r["tempo_total_bypass_fmt"], destaque)
        _dcell(ws, i, 6, r["operadores_resumo"], destaque)

    _autoajustar_colunas(ws, [26, 40, 13, 15, 20, 40])
    ws.freeze_panes = "A2"
    if resumo_ordenado:
        _adicionar_tabela(ws, "ResumoBypass", len(resumo_ordenado) + 1, len(cabecalho))


def _bloco_top_barras(ws, row, titulo, pares):
    """Escreve uma pequena tabela (categoria, valor) a partir de `row` e
    ancora um grafico de barras horizontais ao lado. Retorna a proxima linha
    livre, com espaco suficiente para o grafico nao se sobrepor ao bloco
    seguinte."""
    ws.cell(row=row, column=1, value=titulo).font = Font(bold=True, size=12, color=AZUL)
    row += 1
    inicio_tabela = row
    _hcell(ws, row, 1, "Item")
    _hcell(ws, row, 2, "Valor")
    row += 1
    for categoria, valor in pares:
        _dcell(ws, row, 1, categoria)
        _dcell(ws, row, 2, valor)
        row += 1
    fim_tabela = row - 1

    grafico = BarChart()
    grafico.type = "bar"
    grafico.title = titulo
    dados = Reference(ws, min_col=2, min_row=inicio_tabela, max_row=fim_tabela)
    categorias = Reference(ws, min_col=1, min_row=inicio_tabela + 1, max_row=fim_tabela)
    grafico.add_data(dados, titles_from_data=True)
    grafico.set_categories(categorias)
    grafico.height = max(6, (fim_tabela - inicio_tabela) * 0.6)
    grafico.width = 18
    grafico.legend = None
    ws.add_chart(grafico, f"D{inicio_tabela}")

    linhas_grafico = int(grafico.height / 0.55)
    return max(row, inicio_tabela + linhas_grafico) + 2


def _aba_dashboard(wb, resumo_alarmes, resumo_bypass, sessoes_alarmes, sessoes_bypass, periodo_texto):
    ws = wb.active
    ws.title = "Dashboard"

    ws.cell(row=1, column=1, value="Dashboard - Alarmes e Bypass").font = Font(bold=True, size=16, color=AZUL)
    ws.cell(row=2, column=1, value=periodo_texto).font = Font(italic=True, size=10, color="808080")

    total_atuacoes = sum(r["qtd_atuacoes"] for r in resumo_alarmes)
    total_sozinho = sum(r["qtd_retornou_sozinho"] for r in resumo_alarmes)
    total_reconhecidas = sum(r["qtd_reconhecidas"] for r in resumo_alarmes)
    total_ainda_ativo_a = sum(r["qtd_ainda_ativo"] for r in resumo_alarmes)
    tempo_total_a = sum(r["tempo_total_atuado_s"] for r in resumo_alarmes)

    total_ativacoes_b = sum(r["qtd_ativacoes"] for r in resumo_bypass)
    total_ainda_ativo_b = sum(r["qtd_ainda_ativo"] for r in resumo_bypass)
    tempo_total_b = sum(r["tempo_total_bypass_s"] for r in resumo_bypass)

    kpis = [
        ("Tags de alarme distintas", len(resumo_alarmes)),
        ("Total de atuacoes de alarme", total_atuacoes),
        ("Alarmes retornaram sozinhos", total_sozinho),
        ("Alarmes reconhecidos pelo operador", total_reconhecidas),
        ("Alarmes ainda atuados", total_ainda_ativo_a),
        ("Tempo total atuado (todos os alarmes)", formatar_duracao(tempo_total_a)),
        ("Tags de bypass distintas", len(resumo_bypass)),
        ("Total de ativacoes de bypass", total_ativacoes_b),
        ("Bypass ainda ativos", total_ainda_ativo_b),
        ("Tempo total em bypass (todos)", formatar_duracao(tempo_total_b)),
    ]

    row = 4
    _hcell(ws, row, 1, "Indicador")
    _hcell(ws, row, 2, "Valor")
    row += 1
    for rotulo, valor in kpis:
        _dcell(ws, row, 1, rotulo)
        _dcell(ws, row, 2, valor)
        row += 1
    row += 2

    status = classificar_alarmes.contagem_status(sessoes_alarmes)
    ws.cell(row=row, column=1, value="Status das Atuacoes de Alarme").font = Font(bold=True, size=12, color=AZUL)
    row += 1
    inicio_tabela = row
    _hcell(ws, row, 1, "Status")
    _hcell(ws, row, 2, "Quantidade")
    row += 1
    for rotulo, valor in [
        ("Reconhecido e retornou", status["reconhecidos"]),
        ("Retornou sozinho", status["sozinho"]),
        ("Ainda ativo", status["ainda_ativo"]),
    ]:
        _dcell(ws, row, 1, rotulo)
        _dcell(ws, row, 2, valor)
        row += 1
    fim_tabela = row - 1

    pizza = PieChart()
    pizza.title = "Status das Atuacoes de Alarme"
    dados = Reference(ws, min_col=2, min_row=inicio_tabela, max_row=fim_tabela)
    categorias = Reference(ws, min_col=1, min_row=inicio_tabela + 1, max_row=fim_tabela)
    pizza.add_data(dados, titles_from_data=True)
    pizza.set_categories(categorias)
    pizza.height = 7
    pizza.width = 12
    ws.add_chart(pizza, f"D{inicio_tabela}")
    row += 2

    row = _bloco_top_barras(
        ws, row, "Top 10 Alarmes por Tempo Total Atuado (horas)",
        [(r["tag"], round(r["tempo_total_atuado_s"] / 3600, 1))
         for r in sorted(resumo_alarmes, key=lambda r: -r["tempo_total_atuado_s"])[:10]],
    )

    row = _bloco_top_barras(
        ws, row, "Top 10 Bypass por Tempo Total Ativo (horas)",
        [(r["tag"], round(r["tempo_total_bypass_s"] / 3600, 1))
         for r in sorted(resumo_bypass, key=lambda r: -r["tempo_total_bypass_s"])[:10]],
    )

    por_operador = classificar_bypass.contagem_por_operador(sessoes_bypass)[:10]
    if por_operador:
        row = _bloco_top_barras(ws, row, "Ativacoes de Bypass por Operador", por_operador)

    por_dia = classificar_alarmes.contagem_por_dia(sessoes_alarmes)
    if por_dia:
        ws.cell(row=row, column=1, value="Atuacoes de Alarme por Dia").font = Font(bold=True, size=12, color=AZUL)
        row += 1
        inicio_tabela = row
        _hcell(ws, row, 1, "Data")
        _hcell(ws, row, 2, "Atuacoes")
        row += 1
        for data, qtd in por_dia:
            _dcell(ws, row, 1, data.strftime("%d/%m/%Y"))
            _dcell(ws, row, 2, qtd)
            row += 1
        fim_tabela = row - 1

        tendencia = LineChart()
        tendencia.title = "Atuacoes de Alarme por Dia"
        dados = Reference(ws, min_col=2, min_row=inicio_tabela, max_row=fim_tabela)
        categorias = Reference(ws, min_col=1, min_row=inicio_tabela + 1, max_row=fim_tabela)
        tendencia.add_data(dados, titles_from_data=True)
        tendencia.set_categories(categorias)
        tendencia.height = 8
        tendencia.width = 20
        for serie in tendencia.series:
            serie.marker.symbol = "circle"
            serie.smooth = False
        ws.add_chart(tendencia, f"D{inicio_tabela}")

    _autoajustar_colunas(ws, [34, 14])
    ws.column_dimensions["C"].width = 3


def gerar(caminho_saida, eventos, resumo_alarmes, resumo_bypass, sessoes_alarmes, sessoes_bypass, periodo_texto):
    wb = Workbook()
    _aba_dashboard(wb, resumo_alarmes, resumo_bypass, sessoes_alarmes, sessoes_bypass, periodo_texto)
    _aba_resumo_alarmes(wb, resumo_alarmes)
    _aba_resumo_bypass(wb, resumo_bypass)
    _aba_linha_do_tempo(wb, eventos)
    wb.active = 0
    wb.save(caminho_saida)
